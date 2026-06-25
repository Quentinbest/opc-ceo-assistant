from __future__ import annotations

import argparse
import hashlib
import json
import tempfile
import unicodedata
import xml.etree.ElementTree as ET
import zipfile
from collections.abc import Mapping
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName

CONTRACT_VERSION = "1.0.0"
RESOURCE_ROOT = Path(__file__).parent / "resources"
CONTRACT_PATH = RESOURCE_ROOT / "contracts" / CONTRACT_VERSION / "workbook-contract.json"


def _canonical_value(value: object) -> object:
    if value is None or isinstance(value, (bool, int)):
        return value
    if isinstance(value, float):
        raise TypeError("float values are not permitted in canonical JSON")
    if isinstance(value, str):
        return unicodedata.normalize("NFC", value)
    if isinstance(value, list | tuple):
        return [_canonical_value(item) for item in value]
    if isinstance(value, dict):
        normalized: dict[str, object] = {}
        for key, item in value.items():
            if not isinstance(key, str):
                raise TypeError("canonical JSON object keys must be strings")
            normalized_key = unicodedata.normalize("NFC", key)
            if normalized_key in normalized:
                raise ValueError("duplicate canonical JSON key after NFC normalization")
            normalized[normalized_key] = _canonical_value(item)
        return normalized
    raise TypeError(f"unsupported canonical JSON value: {type(value).__name__}")


def canonical_json_bytes(value: object) -> bytes:
    canonical = _canonical_value(value)
    return (
        json.dumps(
            canonical,
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
            allow_nan=False,
        )
        + "\n"
    ).encode()


def load_contract(path: Path = CONTRACT_PATH) -> dict[str, Any]:
    value = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(value, dict) or value.get("version") != CONTRACT_VERSION:
        raise ValueError("unsupported workbook contract")
    return value


_CONTRACT = load_contract()
EXPECTED_RANGES: Mapping[str, str] = {
    str(tab["sheet"]): str(tab["named_range"]) for tab in _CONTRACT["tabs"]
}


def _record_schema(record_type: str, columns: list[str]) -> dict[str, Any]:
    properties: dict[str, Any] = {name: {"type": ["string", "null"]} for name in columns}
    properties["record_id"] = {
        "type": "string",
        "pattern": rf"^{record_type}_[a-z0-9][a-z0-9_-]{{2,63}}$",
    }
    properties["title"] = {"type": "string", "minLength": 1, "maxLength": 200}
    properties["status"] = {"type": "string", "minLength": 1}
    properties["updated_at"] = {"type": "string", "format": "date-time"}
    return {
        "type": "object",
        "additionalProperties": False,
        "required": ["record_id", "title", "status", "updated_at"],
        "properties": properties,
    }


def build_schema(contract: dict[str, Any]) -> dict[str, Any]:
    definitions = {
        str(tab["type"]): _record_schema(str(tab["type"]), list(tab["columns"]))
        for tab in contract["tabs"]
    }
    definitions.update(
        {
            "brief": {"type": "object", "additionalProperties": False},
            "decision": {"type": "object", "additionalProperties": False},
            "source_snapshot": {"type": "object"},
            "source_projection": {"type": "object"},
            "evidence_reference": {"type": "object"},
            "common_record": {"type": "object"},
        }
    )
    return {
        "$schema": "https://json-schema.org/draft/2020-12/schema",
        "$id": "https://opc.local/schema/opc-workspace/1.0.0",
        "title": "OPC Workspace Schema",
        "type": "object",
        "$defs": definitions,
    }


def _normalize_zip(source: Path, destination: Path) -> None:
    with (
        zipfile.ZipFile(source, "r") as incoming,
        zipfile.ZipFile(
            destination, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9
        ) as outgoing,
    ):
        for name in sorted(incoming.namelist()):
            info = zipfile.ZipInfo(name, date_time=(1980, 1, 1, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            info.external_attr = 0o600 << 16
            content = incoming.read(name)
            if name == "docProps/core.xml":
                namespaces = {
                    "cp": "http://schemas.openxmlformats.org/package/2006/metadata/core-properties",
                    "dc": "http://purl.org/dc/elements/1.1/",
                    "dcterms": "http://purl.org/dc/terms/",
                    "xsi": "http://www.w3.org/2001/XMLSchema-instance",
                }
                for prefix, uri in namespaces.items():
                    ET.register_namespace(prefix, uri)
                root = ET.fromstring(content)
                modified = root.find(f"{{{namespaces['dcterms']}}}modified")
                if modified is None:
                    raise ValueError("workbook core properties lack modified timestamp")
                modified.text = "2000-01-01T00:00:00Z"
                content = ET.tostring(root, encoding="utf-8")
            outgoing.writestr(info, content)


def build_workbook(contract: dict[str, Any], destination: Path) -> None:
    workbook = Workbook()
    active = workbook.active
    if active is None:
        raise ValueError("new workbook has no active worksheet")
    workbook.remove(active)
    workbook.properties.creator = "OPC CEO"
    workbook.properties.lastModifiedBy = "OPC CEO"
    fixed_time = datetime(2000, 1, 1, tzinfo=UTC)
    workbook.properties.created = fixed_time
    workbook.properties.modified = fixed_time
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for tab in contract["tabs"]:
        worksheet = workbook.create_sheet(str(tab["sheet"]))
        columns = list(tab["columns"])
        worksheet.append(columns)
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = f"A1:{worksheet.cell(1, len(columns)).coordinate}"
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
        for index, column in enumerate(columns, 1):
            worksheet.column_dimensions[worksheet.cell(1, index).column_letter].width = max(
                14, min(28, len(column) + 3)
            )
        end = worksheet.cell(1001, len(columns)).coordinate
        reference = f"'{worksheet.title}'!$A$1:${end}"
        workbook.defined_names.add(DefinedName(str(tab["named_range"]), attr_text=reference))

    destination.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory() as temporary:
        raw = Path(temporary) / "raw.xlsx"
        workbook.save(raw)
        _normalize_zip(raw, destination)


def _paths(root: Path) -> dict[str, Path]:
    return {
        "schema": root / "schemas" / CONTRACT_VERSION / "opc-workspace.schema.json",
        "workbook": root / "workbook" / "opc-operating-workbook-v1.xlsx",
        "contract": root / "contracts" / CONTRACT_VERSION / "workbook-contract.json",
    }


def _generate(root: Path) -> dict[str, Path]:
    contract = load_contract()
    paths = _paths(root)
    for path in paths.values():
        path.parent.mkdir(parents=True, exist_ok=True)
    paths["contract"].write_bytes(canonical_json_bytes(contract))
    paths["schema"].write_bytes(canonical_json_bytes(build_schema(contract)))
    build_workbook(contract, paths["workbook"])
    return paths


def generate_resources(root: Path, *, check: bool = False) -> dict[str, Any]:
    if not check:
        result: dict[str, Any] = _generate(root)
        result["ok"] = True
        return result
    expected = _paths(root)
    with tempfile.TemporaryDirectory() as temporary:
        actual = _generate(Path(temporary))
        ok = all(
            expected[name].exists()
            and hashlib.sha256(expected[name].read_bytes()).digest()
            == hashlib.sha256(actual[name].read_bytes()).digest()
            for name in expected
        )
    result = dict(expected)
    result["ok"] = ok
    return result


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("command", choices=["generate"])
    parser.add_argument("--check", action="store_true")
    parser.add_argument("--output", type=Path, default=RESOURCE_ROOT)
    args = parser.parse_args(argv)
    result = generate_resources(args.output, check=args.check)
    return 0 if result["ok"] else 1


if __name__ == "__main__":  # pragma: no cover - exercised by python -m
    raise SystemExit(main())
