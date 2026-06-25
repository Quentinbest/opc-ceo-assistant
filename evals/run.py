from __future__ import annotations

import argparse
import json
import subprocess
import sys
import tempfile
from collections.abc import Sequence
from datetime import UTC, datetime, timedelta
from pathlib import Path
from typing import Any, cast

from openpyxl import load_workbook

from evals.verify_output import verify_case, verify_report
from opc_ceo.briefing import BriefingError, draft_briefing, render_briefing
from opc_ceo.contracts import canonical_json_bytes, generate_resources
from opc_ceo.intake import apply_import, resolve_import, review_import, stage_import
from opc_ceo.workspace import initialize_workspace, secure_replace

ROOT = Path(__file__).resolve().parent
CASES_ROOT = ROOT / "cases"
NOW = datetime(2026, 6, 20, 8, 0, tzinfo=UTC)


def load_cases(root: Path) -> list[dict[str, Any]]:
    cases = [json.loads(path.read_text(encoding="utf-8")) for path in sorted(root.glob("*.json"))]
    return cases


def _metadata(version: str) -> dict[str, Any]:
    value = {"version": version, "modifiedTime": "2026-06-20T00:00:00Z"}
    return {"before": value, "after": dict(value), "spreadsheet_id_hash": "sha256:eval"}


def _refresh_manifest(root: Path, refreshed_at: datetime = NOW) -> None:
    path = root / ".opc" / "manifest.json"
    manifest = json.loads(path.read_text(encoding="utf-8"))
    manifest["source"]["last_successful_refresh_at"] = refreshed_at.isoformat()
    manifest["source"]["last_fully_applied_drive_version"] = "1"
    secure_replace(path, canonical_json_bytes(manifest))


def _seed_priority(root: Path) -> dict[str, Any]:
    record = {
        "type": "priority",
        "record_id": "priority_growth",
        "title": "Grow recurring revenue",
        "status": "active",
        "updated_at": "2026-06-19T08:00:00+08:00",
        "priority_kind": "objective",
        "weight": 5,
        "target_date": "2026-06-21",
        "amount": "30000.00",
        "currency": "CNY",
    }
    secure_replace(
        root / "data" / "priorities" / "priority_growth.json", canonical_json_bytes(record)
    )
    return record


def _workbook(root: Path, *, malformed_receivable: bool = False) -> Path:
    path = cast(Path, generate_resources(root / "generated")["workbook"])
    book = load_workbook(path)
    book["priorities"].append(
        [
            "priority_growth",
            "Grow recurring revenue",
            "active",
            "2026-06-19T08:00:00+08:00",
            None,
            "objective",
            5,
            "2026-06-21",
            "30000.00",
            "CNY",
            "key-account",
            None,
        ]
    )
    if malformed_receivable:
        book["receivables"].append(
            [
                "receivable_acme",
                "June milestone",
                "open",
                "2026-06-19T09:00:00+08:00",
                None,
                "2026-06-18",
                "50000.00",
                "not-a-decimal",
                "CNY",
                False,
                "acme",
                "priority_growth",
            ]
        )
    book.save(path)
    return path


def _normal(root: Path, language: str) -> dict[str, Any]:
    initialize_workspace(root, approved=True)
    record = _seed_priority(root)
    _refresh_manifest(root)
    drafted = draft_briefing(root, now=NOW, language=language)
    view = json.dumps(drafted["top_approval_view"], ensure_ascii=False)
    rendered = render_briefing(
        root,
        drafted["run_id"],
        {"T1": {"kind": "act", "next_action": "Confirm owner and due date"}},
    )
    headings = {
        "zh-CN": "今日 CEO 简报",
        "en": "Daily CEO Brief",
        "bilingual": "Daily CEO Brief / 今日 CEO 简报",
    }
    leaked = record["record_id"] in view or record["title"] in view
    return {
        "assertions": {
            "drafted": drafted["outcome"] == "drafted",
            "language_rendered": headings[language] in rendered["brief_markdown"],
            "opaque_host_view": not leaked,
            "disposition_present": "Disposition: act" in rendered["brief_markdown"],
            "sealed": rendered["outcome"] == "sealed" and bool(rendered["approval_token"]),
        },
        "auto_failures": ["pii_non_top_leakage"] if leaked else [],
    }


def _empty(root: Path) -> dict[str, Any]:
    initialize_workspace(root, approved=True)
    _refresh_manifest(root)
    drafted = draft_briefing(root, now=NOW, language="en")
    rendered = render_briefing(root, drafted["run_id"], {})
    return {
        "assertions": {
            "no_candidates": drafted["outcome"] == "no_candidates",
            "empty_view": drafted["top_approval_view"] == [],
            "empty_render": "No decision-required items." in rendered["brief_markdown"],
            "sealed": rendered["outcome"] == "sealed",
        },
        "auto_failures": [],
    }


def _decline(root: Path) -> dict[str, Any]:
    outcome = initialize_workspace(root, approved=False)
    bypassed = (root / ".opc" / "manifest.json").exists()
    return {
        "assertions": {
            "declined": outcome == "declined",
            "no_workspace_write": not root.exists(),
            "p0b_not_started": not (root / "inbox" / "briefings").exists(),
        },
        "auto_failures": ["bypassed_decline"] if bypassed else [],
    }


def _degraded(root: Path) -> dict[str, Any]:
    initialize_workspace(root, approved=True)
    source = _workbook(root, malformed_receivable=True)
    staged = stage_import(root, source, _metadata("1"), now=NOW)
    review = review_import(root, staged["run_id"])
    return {
        "assertions": {
            "degraded": staged["outcome"] == "staged_degraded",
            "valid_row_retained": review["batch_safe"] == ["priority_growth"],
            "bad_row_quarantined": review["quarantined"] == ["receivable_acme"],
            "no_false_tombstone": review["tombstones"] == [],
        },
        "auto_failures": [],
    }


def _conflict(root: Path) -> dict[str, Any]:
    initialize_workspace(root, approved=True)
    source = _workbook(root)
    first = stage_import(root, source, _metadata("1"), now=NOW)
    sealed = resolve_import(root, first["run_id"], {"batch": "approve", "items": {}})
    apply_import(root, first["run_id"], confirm=sealed["approval_token"], now=NOW)

    book = load_workbook(source)
    book["priorities"].cell(2, 2).value = "Source changed title"
    book.save(source)
    canonical_path = root / "data" / "priorities" / "priority_growth.json"
    canonical = json.loads(canonical_path.read_text(encoding="utf-8"))
    canonical["title"] = "Locally changed title"
    secure_replace(canonical_path, canonical_json_bytes(canonical))
    staged = stage_import(root, source, _metadata("2"), now=NOW)
    review = review_import(root, staged["run_id"])
    return {
        "assertions": {
            "conflict_detected": review["conflicts"] == ["priority_growth"],
            "item_decision_required": review["batch_safe"] == [],
            "prior_canonical_preserved": json.loads(canonical_path.read_text())["title"]
            == "Locally changed title",
        },
        "auto_failures": [],
    }


def _stale(root: Path, *, too_old: bool) -> dict[str, Any]:
    initialize_workspace(root, approved=True)
    _seed_priority(root)
    _refresh_manifest(root)
    offset = timedelta(hours=168, seconds=1) if too_old else timedelta(hours=168)
    drafted = draft_briefing(
        root,
        now=NOW + offset,
        language="en",
        connector_failed=True,
        allow_stale=True,
    )
    if too_old:
        assertions = {
            "diagnostic_only": drafted["outcome"] == "stale_too_old",
            "no_run": drafted["run_id"] is None,
            "no_top": drafted["top_approval_view"] == [],
            "diagnostic_present": drafted["diagnostics"][0]["code"] == "STALE_SOURCE_TOO_OLD",
        }
    else:
        assertions = {
            "boundary_allowed": drafted["outcome"] == "stale_drafted",
            "run_created": isinstance(drafted["run_id"], str),
            "top_present": len(drafted["top_approval_view"]) == 1,
        }
    return {"assertions": assertions, "auto_failures": []}


def _malformed_disposition(root: Path) -> dict[str, Any]:
    initialize_workspace(root, approved=True)
    _seed_priority(root)
    _refresh_manifest(root)
    drafted = draft_briefing(root, now=NOW, language="en")
    blocked = False
    try:
        render_briefing(root, drafted["run_id"], {"T1": {"kind": "act"}})
    except BriefingError:
        blocked = True
    recovered = render_briefing(
        root,
        drafted["run_id"],
        {"T1": {"kind": "act", "next_action": "Corrected action"}},
    )
    return {
        "assertions": {
            "malformed_blocked": blocked,
            "recovery_succeeds": recovered["outcome"] == "sealed",
            "disposition_present": "Corrected action" in recovered["brief_markdown"],
        },
        "auto_failures": [] if blocked else ["missing_disposition"],
    }


def run_case(case: dict[str, Any], workspace: Path) -> dict[str, Any]:
    scenario = case["scenario"]
    if scenario == "normal":
        observed = _normal(workspace, case["language"])
    elif scenario == "empty":
        observed = _empty(workspace)
    elif scenario == "p0a_decline":
        observed = _decline(workspace)
    elif scenario == "degraded_row":
        observed = _degraded(workspace)
    elif scenario == "conflict_review":
        observed = _conflict(workspace)
    elif scenario == "stale_within_limit":
        observed = _stale(workspace, too_old=False)
    elif scenario == "stale_too_old":
        observed = _stale(workspace, too_old=True)
    elif scenario == "malformed_disposition":
        observed = _malformed_disposition(workspace)
    else:
        raise ValueError(f"unknown eval scenario: {scenario}")
    return verify_case(case, observed)


def _run_isolated(case: dict[str, Any], workspace: Path) -> dict[str, Any]:
    completed = subprocess.run(
        [
            sys.executable,
            "-m",
            "evals.run",
            "--case",
            case["id"],
            "--workspace",
            str(workspace),
        ],
        check=True,
        capture_output=True,
        text=True,
    )
    value = json.loads(completed.stdout)
    if not isinstance(value, dict):
        raise ValueError("eval worker returned a non-object")
    return cast(dict[str, Any], value)


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Run OPC CEO deterministic forward evals")
    parser.add_argument("--all", action="store_true")
    parser.add_argument("--case")
    parser.add_argument("--workspace", type=Path)
    parser.add_argument("--output", type=Path)
    parser.add_argument("--model", help="Recorded only; no external model is invoked")
    args = parser.parse_args(argv)
    cases = load_cases(CASES_ROOT)
    if args.case:
        case = next(item for item in cases if item["id"] == args.case)
        if args.workspace is None:
            parser.error("--workspace is required with --case")
        print(json.dumps(run_case(case, args.workspace), ensure_ascii=False, sort_keys=True))
        return 0
    if not args.all:
        parser.error("use --all or --case")
    with tempfile.TemporaryDirectory(prefix="opc-evals-") as temporary:
        root = Path(temporary)
        results = [_run_isolated(case, root / case["id"]) for case in cases]
    report = verify_report(results)
    report["requested_model"] = args.model
    report["model_executed"] = False
    rendered = json.dumps(report, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(rendered, encoding="utf-8")
    print(rendered, end="")
    return 0 if report["passed"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
