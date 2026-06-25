from __future__ import annotations

import hashlib
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from opc_ceo.contracts import EXPECTED_RANGES, canonical_json_bytes, generate_resources


def test_canonical_json_normalizes_unicode_and_rejects_floats() -> None:
    assert canonical_json_bytes({"b": "e\u0301", "a": "x"}) == ('{"a":"x","b":"é"}\n'.encode())
    with pytest.raises(TypeError, match="float"):
        canonical_json_bytes({"amount": 1.5})


def test_contract_generation_is_reproducible_and_complete(tmp_path: Path) -> None:
    first = tmp_path / "first"
    second = tmp_path / "second"

    first_outputs = generate_resources(first)
    second_outputs = generate_resources(second)

    first_workbook = first_outputs["workbook"]
    second_workbook = second_outputs["workbook"]
    assert (
        hashlib.sha256(first_workbook.read_bytes()).digest()
        == hashlib.sha256(second_workbook.read_bytes()).digest()
    )

    workbook = load_workbook(first_workbook, read_only=False, data_only=False)
    assert workbook.sheetnames == list(EXPECTED_RANGES)
    assert set(workbook.defined_names) == set(EXPECTED_RANGES.values())

    schema = json.loads(first_outputs["schema"].read_text())
    assert schema["$schema"] == "https://json-schema.org/draft/2020-12/schema"
    assert {"priority", "pipeline", "receivable", "contract", "project", "risk"} <= set(
        schema["$defs"]
    )
    assert {"brief", "decision", "source_snapshot", "source_projection"} <= set(schema["$defs"])


def test_generate_check_detects_drift(tmp_path: Path) -> None:
    outputs = generate_resources(tmp_path)
    outputs["schema"].write_text("{}\n")

    assert generate_resources(tmp_path, check=True)["ok"] is False
