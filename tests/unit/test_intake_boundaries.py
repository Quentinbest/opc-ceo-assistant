from __future__ import annotations

import json
import os
import shutil
import stat
from datetime import UTC, date, datetime
from decimal import InvalidOperation
from pathlib import Path
from types import SimpleNamespace
from typing import ClassVar

import pytest
from openpyxl import load_workbook

import opc_ceo.intake as intake
from opc_ceo.contracts import canonical_json_bytes, generate_resources
from opc_ceo.workspace import initialize_workspace, secure_replace

NOW = datetime(2026, 6, 20, 8, 0, tzinfo=UTC)


def workspace(tmp_path: Path) -> Path:
    root = tmp_path / "workspace"
    initialize_workspace(root, approved=True)
    return root


def workbook(tmp_path: Path) -> Path:
    source = generate_resources(tmp_path / "generated")["workbook"]
    target = tmp_path / "source.xlsx"
    shutil.copyfile(source, target)
    return target


def metadata(version: str = "1") -> dict[str, object]:
    pair = {"version": version, "modifiedTime": "2026-06-20T00:00:00Z"}
    return {"before": pair, "after": dict(pair)}


def test_read_json_and_metadata_validation(tmp_path: Path) -> None:
    path = tmp_path / "array.json"
    path.write_text("[]\n")
    with pytest.raises(intake.IntakeError, match="expected JSON object"):
        intake._read_json(path)
    invalid_metadata: tuple[object, ...] = (
        None,
        [],
        {"version": 1, "modifiedTime": "x"},
        {"version": "x"},
    )
    for value in invalid_metadata:
        with pytest.raises(intake.IntakeError):
            intake._metadata(value)


@pytest.mark.parametrize(
    ("details", "message"),
    [
        (SimpleNamespace(st_mode=stat.S_IFDIR, st_uid=os.getuid(), st_size=1), "regular"),
        (SimpleNamespace(st_mode=stat.S_IFREG, st_uid=os.getuid() + 1, st_size=1), "owner"),
        (SimpleNamespace(st_mode=stat.S_IFREG, st_uid=os.getuid(), st_size=0), "size"),
        (
            SimpleNamespace(
                st_mode=stat.S_IFREG,
                st_uid=os.getuid(),
                st_size=101 * 1024 * 1024,
            ),
            "size",
        ),
    ],
)
def test_external_artifact_descriptor_validation(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
    details: SimpleNamespace,
    message: str,
) -> None:
    source = tmp_path / "source"
    source.write_bytes(b"x")
    monkeypatch.setattr("opc_ceo.intake.os.fstat", lambda descriptor: details)
    with pytest.raises(intake.IntakeError, match=message):
        intake._copy_external_once(source, tmp_path / "copy")


def test_normalization_boundaries() -> None:
    assert intake._date_string(datetime(2026, 6, 20, tzinfo=UTC)) == "2026-06-20"
    assert intake._date_string(date(2026, 6, 20)) == "2026-06-20"
    assert intake._timestamp(datetime(2026, 6, 20, tzinfo=UTC)).endswith("+00:00")
    with pytest.raises(ValueError, match="offset"):
        intake._timestamp("2026-06-20T08:00:00")
    assert intake._normalize_value("optional", "") is None
    assert intake._normalize_value("amount", "1.20") == "1.20"
    with pytest.raises(InvalidOperation):
        intake._normalize_value("amount", "NaN")
    with pytest.raises(ValueError, match="ISO"):
        intake._normalize_value("currency", "US")
    assert intake._normalize_value("blocked", True) is True
    for value in ("true", "yes", "1"):
        assert intake._normalize_value("blocked", value) is True
    for value in ("false", "no", "0"):
        assert intake._normalize_value("blocked", value) is False
    with pytest.raises(ValueError, match="boolean"):
        intake._normalize_value("blocked", "maybe")
    with pytest.raises(ValueError, match="weight"):
        intake._normalize_value("weight", 6)


def test_defined_range_failures() -> None:
    missing = SimpleNamespace(defined_names={})
    with pytest.raises(intake.IntakeError, match="missing named range"):
        intake._defined_range(missing, "missing")

    class Definition:
        destinations: ClassVar[list[tuple[str, str]]] = [("a", "A1"), ("b", "A1")]

    duplicate = SimpleNamespace(defined_names={"range": Definition()})
    with pytest.raises(intake.IntakeError, match="invalid named range"):
        intake._defined_range(duplicate, "range")


def test_workbook_named_range_header_formula_and_required_failures(tmp_path: Path) -> None:
    source = workbook(tmp_path)
    book = load_workbook(source)
    definition = book.defined_names["opc_priorities_v1"]
    definition.attr_text = "'pipeline'!$A$1:$M$1001"
    book.save(source)
    with pytest.raises(intake.IntakeError, match="points to"):
        intake._normalize_workbook(source)

    source = workbook(tmp_path / "header")
    book = load_workbook(source)
    book["priorities"]["A1"] = "wrong"
    book.save(source)
    with pytest.raises(intake.IntakeError, match="header drift"):
        intake._normalize_workbook(source)

    source = workbook(tmp_path / "formula")
    book = load_workbook(source)
    sheet = book["priorities"]
    sheet.append(
        [
            "priority_formula",
            None,
            "active",
            "2026-06-20T00:00:00Z",
            None,
            "objective",
            "=1+1",
        ]
    )
    book.save(source)
    rows, diagnostics, identity_failure = intake._normalize_workbook(source)
    assert identity_failure is False
    assert rows[0]["quarantined"] is True
    assert {"required", "weight"} <= set(diagnostics[0]["fields"])


def test_amount_anomaly_all_boundaries() -> None:
    assert intake._amount_anomaly({}, {"type": "risk"}, {}) is False
    assert (
        intake._amount_anomaly({"currency": "USD"}, {"type": "priority", "currency": "CNY"}, {})
        is True
    )
    assert (
        intake._amount_anomaly({"currency": None}, {"type": "priority", "currency": None}, {})
        is False
    )
    with pytest.raises(intake.IntakeError, match="threshold"):
        intake._amount_anomaly(
            {"currency": "EUR"},
            {"type": "priority", "currency": "EUR", "amount": "1"},
            {},
        )
    thresholds = {"CNY": "100"}
    assert (
        intake._amount_anomaly(
            {"currency": "CNY", "amount": "0"},
            {"type": "priority", "currency": "CNY", "amount": "101"},
            thresholds,
        )
        is True
    )
    assert (
        intake._amount_anomaly(
            {"currency": "CNY", "amount": "100"},
            {"type": "priority", "currency": "CNY", "amount": "301"},
            thresholds,
        )
        is True
    )
    assert (
        intake._amount_anomaly(
            {"currency": "CNY", "amount": "100"},
            {"type": "priority", "currency": "CNY", "amount": "150"},
            thresholds,
        )
        is False
    )
    assert (
        intake._amount_anomaly(
            {
                "currency": "CNY",
                "total_amount": "100",
                "outstanding_amount": "100",
            },
            {
                "type": "receivable",
                "currency": "CNY",
                "total_amount": "150",
                "outstanding_amount": "301",
            },
            thresholds,
        )
        is True
    )
    assert (
        intake._amount_anomaly(
            {
                "currency": "CNY",
                "total_amount": "0",
                "outstanding_amount": "0",
            },
            {
                "type": "receivable",
                "currency": "CNY",
                "total_amount": "50",
                "outstanding_amount": "50",
            },
            thresholds,
        )
        is False
    )


def test_stage_invalid_metadata_unchanged_collision_and_bad_source(tmp_path: Path) -> None:
    root = workspace(tmp_path)
    source = workbook(tmp_path)
    blocked = intake.stage_import(root, source, {}, now=NOW)
    assert blocked["diagnostics"][0]["code"] == "CONNECTOR_METADATA_ERROR"

    manifest_path = root / ".opc" / "manifest.json"
    manifest = json.loads(manifest_path.read_text())
    manifest["source"]["last_fully_applied_drive_version"] = "1"
    secure_replace(manifest_path, canonical_json_bytes(manifest))
    assert intake.stage_import(root, source, metadata(), now=NOW)["outcome"] == "unchanged"

    manifest["source"]["last_fully_applied_drive_version"] = None
    secure_replace(manifest_path, canonical_json_bytes(manifest))
    first = intake.stage_import(root, source, metadata(), now=NOW)
    second = intake.stage_import(root, source, metadata(), now=NOW)
    assert first["run_id"] != second["run_id"]

    bad = tmp_path / "bad.xlsx"
    bad.write_bytes(b"not xlsx")
    assert intake.stage_import(root, bad, metadata("2"), now=NOW)["outcome"] == "blocked"


def test_resolution_and_seal_failures(tmp_path: Path) -> None:
    root = workspace(tmp_path)
    source = workbook(tmp_path)
    staged = intake.stage_import(root, source, metadata(), now=NOW)
    run_id = staged["run_id"]
    assert isinstance(run_id, str)
    with pytest.raises(intake.IntakeError, match="resolution requires"):
        intake.resolve_import(root, run_id, {"batch": "invalid", "items": []})

    run_root = root / "inbox" / "imports" / run_id
    (run_root / "diff.json").write_bytes(
        canonical_json_bytes(
            [{"record_id": "risk_new", "type": "risk", "status": "add", "source_hash": "x"}]
        )
    )
    with pytest.raises(intake.IntakeError, match="invalid item decision"):
        intake.resolve_import(root, run_id, {"batch": "approve", "items": {"risk_new": "invalid"}})

    (run_root / "diff.json").write_bytes(canonical_json_bytes([]))
    sealed = intake.resolve_import(root, run_id, {"batch": "approve", "items": {}})
    (run_root / "envelope.json").write_text("{}\n")
    with pytest.raises(intake.ApprovalMismatch, match="sealed files"):
        intake.apply_import(root, run_id, confirm=sealed["approval_token"], now=NOW)


def test_apply_clears_pending_and_missing_tombstone_target(tmp_path: Path) -> None:
    root = workspace(tmp_path)
    source = workbook(tmp_path)
    staged = intake.stage_import(root, source, metadata(), now=NOW)
    run_id = staged["run_id"]
    assert isinstance(run_id, str)
    run_root = root / "inbox" / "imports" / run_id
    projection = {
        "record_id": "risk_existing",
        "applied": {"source_hash": "same", "canonical_hash": None},
        "pending": {"status": "rejected"},
    }
    secure_replace(
        root / "data" / "source_projection" / "risk_existing.json",
        canonical_json_bytes(projection),
    )
    (run_root / "diff.json").write_bytes(
        canonical_json_bytes(
            [
                {
                    "record_id": "risk_existing",
                    "type": "risk",
                    "status": "unchanged",
                    "source_hash": "same",
                }
            ]
        )
    )
    sealed = intake.resolve_import(root, run_id, {"batch": "approve", "items": {}})
    intake.apply_import(root, run_id, confirm=sealed["approval_token"], now=NOW)
    cleared = json.loads((root / "data" / "source_projection" / "risk_existing.json").read_text())
    assert cleared["pending"] is None

    review_run = root / "inbox" / "imports" / "review_unknown"
    review_run.mkdir()
    (review_run / "diff.json").write_bytes(
        canonical_json_bytes([{"record_id": "risk_unknown", "type": "risk", "status": "custom"}])
    )
    assert all(not values for values in intake.review_import(root, "review_unknown").values())

    no_pending_run = root / "inbox" / "imports" / "no_pending"
    shutil.copytree(run_root, no_pending_run)
    for name in ("resolution.json", "apply_plan.json", "seal.json", "apply_result.json"):
        (no_pending_run / name).unlink(missing_ok=True)
    no_pending_projection = json.loads(
        (root / "data" / "source_projection" / "risk_existing.json").read_text()
    )
    assert no_pending_projection["pending"] is None
    (no_pending_run / "diff.json").write_bytes(
        canonical_json_bytes(
            [
                {
                    "record_id": "risk_existing",
                    "type": "risk",
                    "status": "unchanged",
                    "source_hash": "same",
                }
            ]
        )
    )
    sealed_no_pending = intake.resolve_import(root, "no_pending", {"batch": "approve", "items": {}})
    intake.apply_import(
        root,
        "no_pending",
        confirm=sealed_no_pending["approval_token"],
        now=NOW,
    )

    root2 = workspace(tmp_path / "second")
    source2 = workbook(tmp_path / "second")
    staged2 = intake.stage_import(root2, source2, metadata(), now=NOW)
    run2 = staged2["run_id"]
    assert isinstance(run2, str)
    run_root2 = root2 / "inbox" / "imports" / run2
    (run_root2 / "diff.json").write_bytes(
        canonical_json_bytes(
            [
                {
                    "record_id": "risk_missing",
                    "type": "risk",
                    "status": "tombstone_pending",
                    "source_hash": "x",
                }
            ]
        )
    )
    sealed2 = intake.resolve_import(
        root2,
        run2,
        {"batch": "approve", "items": {"risk_missing": "approve"}},
    )
    with pytest.raises(intake.IntakeError, match="missing canonical tombstone"):
        intake.apply_import(root2, run2, confirm=sealed2["approval_token"], now=NOW)
