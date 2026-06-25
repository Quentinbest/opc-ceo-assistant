from __future__ import annotations

import ast
import json
import shutil
import subprocess
import sys
from datetime import UTC, datetime, timedelta
from pathlib import Path

import pytest
from openpyxl import load_workbook

import opc_ceo.briefing as briefing
import opc_ceo.installer as installer
import opc_ceo.intake as intake
from opc_ceo.contracts import canonical_json_bytes, generate_resources
from opc_ceo.diagnostics import workspace_status
from opc_ceo.workspace import (
    FaultInjected,
    FaultPoint,
    WorkspaceError,
    fault_checkpoint,
    initialize_workspace,
    secure_replace,
    workspace_lock,
)

NOW = datetime(2026, 6, 20, 8, 0, tzinfo=UTC)


def test_fault_registry_matches_workspace_inventory(tmp_path: Path) -> None:
    root = tmp_path / "workspace"
    initialize_workspace(root, approved=True)
    inventory = json.loads((root / ".opc" / "write-boundaries.json").read_text())

    assert {item["fault_point"] for item in inventory} == {point.value for point in FaultPoint}
    assert all({"fault_point", "owner", "purpose"} == set(item) for item in inventory)


@pytest.mark.parametrize(
    ("module", "prefix"),
    [(intake, "import."), (briefing, "brief."), (installer, "install.")],
)
def test_ast_fault_guard_matches_each_critical_write_registry(module: object, prefix: str) -> None:
    tree = ast.parse(Path(module.__file__).read_text(encoding="utf-8"))  # type: ignore[attr-defined]
    referenced = {
        node.attr
        for node in ast.walk(tree)
        if isinstance(node, ast.Attribute)
        and isinstance(node.value, ast.Name)
        and node.value.id == "FaultPoint"
    }
    expected = {point.name for point in FaultPoint if point.value.startswith(prefix)}
    assert referenced == expected


@pytest.mark.parametrize("point", list(FaultPoint))
def test_every_fault_point_is_injectable(
    monkeypatch: pytest.MonkeyPatch, point: FaultPoint
) -> None:
    monkeypatch.setenv("OPC_FAULT_POINT", point.value)
    with pytest.raises(FaultInjected, match=point.value):
        fault_checkpoint(point)


def test_non_matching_fault_point_is_a_noop(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("OPC_FAULT_POINT", "different")
    fault_checkpoint(FaultPoint.IMPORT_APPLY_STARTED)


def test_import_apply_visits_every_import_checkpoint(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    root = tmp_path / "workspace"
    initialize_workspace(root, approved=True)
    generated = generate_resources(tmp_path / "generated")["workbook"]
    source = tmp_path / "source.xlsx"
    shutil.copyfile(generated, source)
    book = load_workbook(source)
    book["priorities"].append(
        [
            "priority_one",
            "Priority",
            "active",
            "2026-06-20T00:00:00Z",
            None,
            "objective",
            1,
        ]
    )
    book.save(source)
    pair = {"version": "1", "modifiedTime": "2026-06-20T00:00:00Z"}
    staged = intake.stage_import(root, source, {"before": pair, "after": pair}, now=NOW)
    run_id = staged["run_id"]
    assert isinstance(run_id, str)
    sealed = intake.resolve_import(root, run_id, {"batch": "approve", "items": {}})
    visited: list[FaultPoint] = []
    monkeypatch.setattr(intake, "fault_checkpoint", visited.append)

    intake.apply_import(root, run_id, confirm=sealed["approval_token"], now=NOW)

    assert set(visited) == {point for point in FaultPoint if point.value.startswith("import.")}


def test_brief_apply_visits_every_brief_checkpoint(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    root = tmp_path / "workspace"
    initialize_workspace(root, approved=True)
    manifest_path = root / ".opc" / "manifest.json"
    manifest = json.loads(manifest_path.read_text())
    manifest["source"]["last_successful_refresh_at"] = NOW.isoformat()
    manifest["source"]["last_fully_applied_drive_version"] = "1"
    secure_replace(manifest_path, canonical_json_bytes(manifest))
    secure_replace(
        root / "data" / "risks" / "risk_one.json",
        canonical_json_bytes(
            {
                "type": "risk",
                "record_id": "risk_one",
                "title": "Risk",
                "status": "active",
                "updated_at": "2026-06-20T00:00:00Z",
                "severity": "high",
            }
        ),
    )
    drafted = briefing.draft_briefing(root, now=NOW, language="en")
    run_id = drafted["run_id"]
    assert isinstance(run_id, str)
    rendered = briefing.render_briefing(
        root, run_id, {"T1": {"kind": "dismiss", "reason": "accepted"}}
    )
    visited: list[FaultPoint] = []
    monkeypatch.setattr(briefing, "fault_checkpoint", visited.append)

    briefing.apply_briefing(root, run_id, confirm=rendered["approval_token"], now=NOW)

    assert set(visited) == {point for point in FaultPoint if point.value.startswith("brief.")}


def test_installer_update_visits_every_install_checkpoint(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    project = tmp_path / "project"
    source = project / "skills" / "opc-ceo-office"
    source.mkdir(parents=True)
    (source / "SKILL.md").write_text("v1")
    home = tmp_path / "codex"
    installer.install_skill(project, home, install_tool=False)
    (source / "SKILL.md").write_text("v2")
    visited: list[FaultPoint] = []
    monkeypatch.setattr(installer, "fault_checkpoint", visited.append)
    monkeypatch.setattr("opc_ceo.installer.subprocess.run", lambda *args, **kwargs: None)

    installer.install_skill(project, home, update=True, install_tool=True)

    assert set(visited) == {point for point in FaultPoint if point.value.startswith("install.")}


def test_workspace_lock_rejects_a_second_process(tmp_path: Path) -> None:
    root = tmp_path / "workspace"
    initialize_workspace(root, approved=True)
    script = (
        "import sys,time; from pathlib import Path; "
        "from opc_ceo.workspace import workspace_lock; "
        "root=Path(sys.argv[1]); "
        "ctx=workspace_lock(root,'process'); ctx.__enter__(); "
        "print('locked', flush=True); time.sleep(5); ctx.__exit__(None,None,None)"
    )
    process = subprocess.Popen(
        [sys.executable, "-c", script, str(root)],
        stdout=subprocess.PIPE,
        text=True,
    )
    try:
        assert process.stdout is not None
        assert process.stdout.readline().strip() == "locked"
        with (
            pytest.raises(WorkspaceError, match="active Workspace lock"),
            workspace_lock(root, "process"),
        ):
            pass
    finally:
        process.terminate()
        process.wait(timeout=5)


def _sealed_import(tmp_path: Path) -> tuple[Path, str, str]:
    root = tmp_path / "workspace"
    initialize_workspace(root, approved=True)
    source = generate_resources(tmp_path / "generated")["workbook"]
    book = load_workbook(source)
    book["priorities"].append(
        ["priority_one", "Priority", "active", "2026-06-20T00:00:00Z", None, "objective", 1]
    )
    book.save(source)
    pair = {"version": "1", "modifiedTime": "2026-06-20T00:00:00Z"}
    staged = intake.stage_import(root, source, {"before": pair, "after": pair}, now=NOW)
    run_id = staged["run_id"]
    sealed = intake.resolve_import(root, run_id, {"batch": "approve", "items": {}})
    return root, run_id, sealed["approval_token"]


@pytest.mark.parametrize(
    "point", [point for point in FaultPoint if point.value.startswith("import.")]
)
def test_import_apply_resumes_after_every_fault(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path, point: FaultPoint
) -> None:
    root, run_id, token = _sealed_import(tmp_path)
    monkeypatch.setenv("OPC_FAULT_POINT", point.value)
    with pytest.raises(FaultInjected):
        intake.apply_import(root, run_id, confirm=token, now=NOW)
    monkeypatch.delenv("OPC_FAULT_POINT")

    outcome = intake.apply_import(root, run_id, confirm=token, now=NOW + timedelta(hours=1))

    assert outcome["outcome"] in {"applied", "already_applied"}
    assert workspace_status(root)["outcome"] == "healthy"
    assert len(list((root / "data" / "source_snapshots").glob("*.json"))) == 1


def _sealed_brief(tmp_path: Path) -> tuple[Path, str, str]:
    root = tmp_path / "workspace"
    initialize_workspace(root, approved=True)
    manifest_path = root / ".opc" / "manifest.json"
    manifest = json.loads(manifest_path.read_text())
    manifest["source"]["last_successful_refresh_at"] = NOW.isoformat()
    manifest["source"]["last_fully_applied_drive_version"] = "1"
    secure_replace(manifest_path, canonical_json_bytes(manifest))
    secure_replace(
        root / "data" / "risks" / "risk_one.json",
        canonical_json_bytes(
            {
                "type": "risk",
                "record_id": "risk_one",
                "title": "Risk",
                "status": "active",
                "updated_at": "2026-06-20T00:00:00Z",
                "severity": "high",
            }
        ),
    )
    drafted = briefing.draft_briefing(root, now=NOW, language="en")
    rendered = briefing.render_briefing(
        root, drafted["run_id"], {"T1": {"kind": "dismiss", "reason": "accepted"}}
    )
    return root, drafted["run_id"], rendered["approval_token"]


@pytest.mark.parametrize(
    "point", [point for point in FaultPoint if point.value.startswith("brief.")]
)
def test_brief_apply_resumes_after_every_fault(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path, point: FaultPoint
) -> None:
    root, run_id, token = _sealed_brief(tmp_path)
    monkeypatch.setenv("OPC_FAULT_POINT", point.value)
    with pytest.raises(FaultInjected):
        briefing.apply_briefing(root, run_id, confirm=token, now=NOW)
    monkeypatch.delenv("OPC_FAULT_POINT")

    outcome = briefing.apply_briefing(root, run_id, confirm=token, now=NOW + timedelta(hours=1))

    assert outcome["outcome"] in {"applied", "already_applied"}
    assert workspace_status(root)["outcome"] == "healthy"
    assert len(list((root / "data" / "briefs").glob("*_r*.json"))) == 1


@pytest.mark.parametrize(
    "point", [point for point in FaultPoint if point.value.startswith("install.")]
)
def test_installer_retries_after_every_fault(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path, point: FaultPoint
) -> None:
    project = tmp_path / "project"
    source = project / "skills" / "opc-ceo-office"
    source.mkdir(parents=True)
    (source / "SKILL.md").write_text("v1")
    home = tmp_path / "codex"
    installer.install_skill(project, home, install_tool=False)
    (source / "SKILL.md").write_text("v2")
    monkeypatch.setattr("opc_ceo.installer.subprocess.run", lambda *args, **kwargs: None)
    monkeypatch.setenv("OPC_FAULT_POINT", point.value)
    with pytest.raises(FaultInjected):
        installer.install_skill(project, home, update=True, install_tool=True)
    monkeypatch.delenv("OPC_FAULT_POINT")

    installer.install_skill(project, home, update=True, install_tool=True)

    assert (
        installer.install_skill(project, home, check=True, install_tool=False)["outcome"]
        == "up_to_date"
    )
    assert (home / "skills" / "opc-ceo-office" / "SKILL.md").read_text() == "v2"


def test_apply_skips_duplicate_completion_audit(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    import_root, import_run, import_token = _sealed_import(tmp_path / "import")
    monkeypatch.setattr(
        intake,
        "event_exists",
        lambda root, event, kind, run_id: event == "apply_completed",
    )
    assert (
        intake.apply_import(import_root, import_run, confirm=import_token, now=NOW)["outcome"]
        == "applied"
    )

    brief_root, brief_run, brief_token = _sealed_brief(tmp_path / "brief")
    monkeypatch.setattr(
        briefing,
        "event_exists",
        lambda root, event, kind, run_id: event == "apply_completed",
    )
    assert (
        briefing.apply_briefing(brief_root, brief_run, confirm=brief_token, now=NOW)["outcome"]
        == "applied"
    )
