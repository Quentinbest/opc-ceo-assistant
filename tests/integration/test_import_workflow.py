from __future__ import annotations

import json
import shutil
from datetime import UTC, datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

from opc_ceo.contracts import generate_resources
from opc_ceo.diagnostics import workspace_status
from opc_ceo.intake import (
    ApprovalMismatch,
    IntakeError,
    apply_import,
    resolve_import,
    review_import,
    stage_import,
)
from opc_ceo.workspace import initialize_workspace, secure_replace

NOW = datetime(2026, 6, 20, 8, 0, tzinfo=UTC)


def metadata(version: str = "3") -> dict[str, object]:
    value = {"version": version, "modifiedTime": "2026-06-20T00:00:00Z"}
    return {"before": value, "after": dict(value), "spreadsheet_id_hash": "sha256:probe"}


def workbook_fixture(tmp_path: Path, *, bad_amount: bool = False, duplicate: bool = False) -> Path:
    generated = generate_resources(tmp_path / "generated")["workbook"]
    path = tmp_path / "source.xlsx"
    shutil.copyfile(generated, path)
    workbook = load_workbook(path)
    priority = workbook["priorities"]
    priority.append(
        [
            "priority_growth",
            "Grow recurring revenue",
            "active",
            "2026-06-19T08:00:00+08:00",
            None,
            "objective",
            5,
            "2026-06-21",
            "30000.00",
            "CNY",
            "key-account",
            None,
        ]
    )
    if duplicate:
        priority.append(next(priority.iter_rows(min_row=2, max_row=2, values_only=True)))
    receivables = workbook["receivables"]
    receivables.append(
        [
            "receivable_acme",
            "June milestone",
            "open",
            "2026-06-19T09:00:00+08:00",
            None,
            "2026-06-18",
            "50000.00",
            "not-a-decimal" if bad_amount else "12000.00",
            "CNY",
            False,
            "acme",
            "priority_growth",
        ]
    )
    workbook.save(path)
    return path


def initialized_workspace(tmp_path: Path) -> Path:
    root = tmp_path / "workspace"
    assert initialize_workspace(root, approved=True) == "initialized"
    return root


def apply_initial(root: Path, source: Path) -> str:
    staged = stage_import(root, source, metadata(), now=NOW)
    run_id = staged["run_id"]
    assert isinstance(run_id, str)
    sealed = resolve_import(root, run_id, {"batch": "approve", "items": {}})
    apply_import(root, run_id, confirm=sealed["approval_token"], now=NOW)
    return run_id


def test_stage_resolve_apply_is_sealed_and_idempotent(tmp_path: Path) -> None:
    root = initialized_workspace(tmp_path)
    source = workbook_fixture(tmp_path)

    staged = stage_import(root, source, metadata(), now=NOW)
    assert staged["outcome"] == "staged_clean"
    run_id = staged["run_id"]
    run_root = root / "inbox" / "imports" / run_id
    assert not (run_root / "source.xlsx").exists()
    assert (run_root / "normalized.jsonl").exists()

    review = review_import(root, run_id)
    assert review["batch_safe"] == ["priority_growth", "receivable_acme"]
    assert review["quarantined"] == []

    sealed = resolve_import(root, run_id, {"batch": "approve", "items": {}})
    assert sealed["outcome"] == "sealed"
    token = sealed["approval_token"]

    with pytest.raises(ApprovalMismatch):
        apply_import(root, run_id, confirm=f"{run_id}:wrong", now=NOW)

    applied = apply_import(root, run_id, confirm=token, now=NOW)
    assert applied["outcome"] == "applied"
    assert (root / "data" / "priorities" / "priority_growth.json").exists()
    projection = json.loads(
        (root / "data" / "source_projection" / "priority_growth.json").read_text()
    )
    assert projection["applied"]["drive_version"] == "3"
    assert projection["pending"] is None
    assert apply_import(root, run_id, confirm=token, now=NOW)["outcome"] == "already_applied"
    status = workspace_status(root)
    assert status["imports"]["runs"] == 1
    assert status["imports"]["states"]["applied"] == 1
    assert status["imports"]["latest"]["drive_version"] == "3"
    assert status["imports"]["latest"]["state"] == "applied"


def test_stage_blocks_metadata_drift_and_duplicate_identity(tmp_path: Path) -> None:
    root = initialized_workspace(tmp_path)
    source = workbook_fixture(tmp_path, duplicate=True)
    drifted = metadata()
    drifted["after"] = {"version": "4", "modifiedTime": "2026-06-20T00:01:00Z"}

    assert stage_import(root, source, drifted, now=NOW)["outcome"] == "blocked"
    blocked = stage_import(root, source, metadata(), now=NOW)
    assert blocked["outcome"] == "blocked"
    assert {item["code"] for item in blocked["diagnostics"]} == {"IDENTITY_CONTRACT_ERROR"}


def test_valid_id_bad_row_is_quarantined_without_tombstones(tmp_path: Path) -> None:
    root = initialized_workspace(tmp_path)
    source = workbook_fixture(tmp_path, bad_amount=True)

    staged = stage_import(root, source, metadata(), now=NOW)
    assert staged["outcome"] == "staged_degraded"
    review = review_import(root, staged["run_id"])
    assert review["batch_safe"] == ["priority_growth"]
    assert review["quarantined"] == ["receivable_acme"]
    assert review["tombstones"] == []


def test_amount_anomaly_and_three_way_conflict_require_item_decisions(tmp_path: Path) -> None:
    root = initialized_workspace(tmp_path)
    source = workbook_fixture(tmp_path)
    apply_initial(root, source)

    workbook = load_workbook(source)
    workbook["receivables"].cell(2, 8).value = "50000.00"
    workbook["priorities"].cell(2, 2).value = "Source changed title"
    workbook.save(source)
    canonical_path = root / "data" / "priorities" / "priority_growth.json"
    canonical = json.loads(canonical_path.read_text())
    canonical["title"] = "Locally changed title"
    secure_replace(canonical_path, json.dumps(canonical, sort_keys=True).encode() + b"\n")

    staged = stage_import(root, source, metadata("4"), now=NOW)
    review = review_import(root, staged["run_id"])
    assert review["anomalies"] == ["receivable_acme"]
    assert review["conflicts"] == ["priority_growth"]
    with pytest.raises(IntakeError, match="item decision required"):
        resolve_import(root, staged["run_id"], {"batch": "approve", "items": {}})


def test_complete_snapshot_tombstone_and_rejected_pending(tmp_path: Path) -> None:
    root = initialized_workspace(tmp_path)
    source = workbook_fixture(tmp_path)
    apply_initial(root, source)

    workbook = load_workbook(source)
    for cell in workbook["priorities"][2]:
        cell.value = None
    workbook["receivables"].cell(2, 2).value = "Changed but rejected"
    workbook.save(source)

    staged = stage_import(root, source, metadata("4"), now=NOW)
    review = review_import(root, staged["run_id"])
    assert review["tombstones"] == ["priority_growth"]
    sealed = resolve_import(
        root,
        staged["run_id"],
        {
            "batch": "reject",
            "items": {"priority_growth": "approve"},
        },
    )
    apply_import(root, staged["run_id"], confirm=sealed["approval_token"], now=NOW)
    archived = json.loads((root / "data" / "priorities" / "priority_growth.json").read_text())
    assert archived["status"] == "archived"
    pending = json.loads((root / "data" / "source_projection" / "receivable_acme.json").read_text())
    assert pending["pending"]["status"] == "rejected"
    assert pending["applied"]["drive_version"] == "3"


def test_apply_rechecks_projection_preconditions(tmp_path: Path) -> None:
    root = initialized_workspace(tmp_path)
    source = workbook_fixture(tmp_path)
    apply_initial(root, source)
    workbook = load_workbook(source)
    workbook["priorities"].cell(2, 2).value = "A newer source value"
    workbook.save(source)
    staged = stage_import(root, source, metadata("4"), now=NOW)
    sealed = resolve_import(root, staged["run_id"], {"batch": "approve", "items": {}})

    projection_path = root / "data" / "source_projection" / "priority_growth.json"
    projection = json.loads(projection_path.read_text())
    projection["pending"] = {"status": "external-change"}
    secure_replace(projection_path, json.dumps(projection, sort_keys=True).encode() + b"\n")

    with pytest.raises(ApprovalMismatch, match="precondition"):
        apply_import(
            root,
            staged["run_id"],
            confirm=sealed["approval_token"],
            now=NOW,
        )
