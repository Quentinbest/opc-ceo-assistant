"""Validate the Connector-to-Python spreadsheet artifact contract.

This probe is deliberately independent from the production package. It records
metadata, file properties, hashes, named-range presence, and cleanup semantics;
it never serializes cell values.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import stat
import unicodedata
from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

PROBE_VERSION = "1.0.0"
EXPECTED_RANGES: Mapping[str, str] = {
    "opc_priorities_v1": "priorities",
    "opc_pipeline_v1": "pipeline",
    "opc_receivables_v1": "receivables",
    "opc_contracts_v1": "contracts",
    "opc_projects_v1": "projects",
    "opc_risks_v1": "risks",
}


class CapabilityError(ValueError):
    """Raised when the Connector handoff does not satisfy the gate."""


@dataclass(frozen=True)
class ExportEvidence:
    absolute_path: str
    owner_uid: int
    size_bytes: int
    sha256: str
    named_ranges: tuple[str, ...]
    sheets: tuple[str, ...]

    def as_receipt(self) -> dict[str, Any]:
        return {
            "absolute_local_path": True,
            "regular_file": True,
            "owner_uid": self.owner_uid,
            "size_bytes": self.size_bytes,
            "sha256": self.sha256,
            "python_read": {
                "library": "openpyxl",
                "read_only": True,
                "data_only": False,
                "succeeded": True,
            },
            "named_ranges": list(self.named_ranges),
            "sheet_names": list(self.sheets),
        }


def canonical_json_bytes(value: Any) -> bytes:
    normalized = _normalize_strings(value)
    text = json.dumps(
        normalized,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        allow_nan=False,
    )
    return (text + "\n").encode("utf-8")


def _normalize_strings(value: Any) -> Any:
    if isinstance(value, str):
        return unicodedata.normalize("NFC", value)
    if isinstance(value, list):
        return [_normalize_strings(item) for item in value]
    if isinstance(value, tuple):
        return [_normalize_strings(item) for item in value]
    if isinstance(value, dict):
        return {_normalize_strings(key): _normalize_strings(item) for key, item in value.items()}
    return value


def _read_regular_file(path: Path) -> tuple[bytes, os.stat_result]:
    if not path.is_absolute():
        raise CapabilityError("export path must be absolute")
    flags = os.O_RDONLY
    if hasattr(os, "O_NOFOLLOW"):
        flags |= os.O_NOFOLLOW
    try:
        descriptor = os.open(path, flags)
    except OSError as error:
        raise CapabilityError(f"cannot safely open export: {error}") from error
    try:
        metadata = os.fstat(descriptor)
        if not stat.S_ISREG(metadata.st_mode):
            raise CapabilityError("export is not a regular file")
        if metadata.st_uid != os.getuid():
            raise CapabilityError("export is not owned by the current user")
        chunks: list[bytes] = []
        while chunk := os.read(descriptor, 1024 * 1024):
            chunks.append(chunk)
        return b"".join(chunks), metadata
    finally:
        os.close(descriptor)


def validate_export(
    export_path: str | Path,
    expected_ranges: Mapping[str, str] = EXPECTED_RANGES,
) -> ExportEvidence:
    path = Path(export_path)
    content, metadata = _read_regular_file(path)
    if not content:
        raise CapabilityError("export is empty")

    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=False)
    except Exception as error:  # openpyxl exposes multiple parse exception types
        raise CapabilityError(f"openpyxl could not read export: {error}") from error

    try:
        sheets = tuple(workbook.sheetnames)
        available_names = set(workbook.defined_names)
        missing_names = sorted(set(expected_ranges) - available_names)
        if missing_names:
            raise CapabilityError("missing named ranges: " + ", ".join(missing_names))

        wrong_destinations: list[str] = []
        for range_name, expected_sheet in expected_ranges.items():
            definition = workbook.defined_names[range_name]
            destinations = tuple(definition.destinations)
            if not any(sheet_name == expected_sheet for sheet_name, _ in destinations):
                wrong_destinations.append(range_name)
        if wrong_destinations:
            raise CapabilityError(
                "named ranges point to unexpected sheets: " + ", ".join(sorted(wrong_destinations))
            )
        missing_sheets = sorted(set(expected_ranges.values()) - set(sheets))
        if missing_sheets:
            raise CapabilityError("missing sheets: " + ", ".join(missing_sheets))
    finally:
        workbook.close()

    return ExportEvidence(
        absolute_path=str(path.resolve(strict=True)),
        owner_uid=metadata.st_uid,
        size_bytes=metadata.st_size,
        sha256=hashlib.sha256(content).hexdigest(),
        named_ranges=tuple(sorted(expected_ranges)),
        sheets=tuple(expected_ranges.values()),
    )


def validate_metadata_pair(before: Mapping[str, Any], after: Mapping[str, Any]) -> dict[str, str]:
    required = ("version", "modifiedTime")
    for label, metadata in (("before", before), ("after", after)):
        missing = [key for key in required if not metadata.get(key)]
        if missing:
            raise CapabilityError(f"{label} metadata missing: " + ", ".join(sorted(missing)))
    if str(before["version"]) != str(after["version"]):
        raise CapabilityError("Drive version changed during export")
    if str(before["modifiedTime"]) != str(after["modifiedTime"]):
        raise CapabilityError("Drive modifiedTime changed during export")
    return {
        "version": str(after["version"]),
        "modifiedTime": str(after["modifiedTime"]),
    }


def validate_cleanup_contract(contract: Mapping[str, Any]) -> dict[str, Any]:
    if contract.get("owner") != "host":
        raise CapabilityError("Connector attachment owner must be host")
    automatic = contract.get("automatic_cleanup") is True
    disposable = contract.get("safe_disposable_path") is True
    if not (automatic or disposable):
        raise CapabilityError("host must guarantee automatic cleanup or a safe disposable path")
    return {
        "owner": "host",
        "automatic_cleanup": automatic,
        "safe_disposable_path": disposable,
        "contract_source": str(contract.get("contract_source", "")),
    }


def passing_receipt(
    metadata_before: Mapping[str, Any],
    metadata_after: Mapping[str, Any],
    export: ExportEvidence,
    cleanup_contract: Mapping[str, Any],
) -> dict[str, Any]:
    return {
        "receipt_schema_version": PROBE_VERSION,
        "status": "passed",
        "capabilities": {
            "drive_metadata": validate_metadata_pair(metadata_before, metadata_after),
            "xlsx_export": export.as_receipt(),
            "host_attachment_cleanup": validate_cleanup_contract(cleanup_contract),
            "raw_cells_entered_host_context": False,
        },
    }


def failure_receipt(
    *, phase: str, error_code: str, message: str, details: Mapping[str, Any]
) -> dict[str, Any]:
    return {
        "receipt_schema_version": PROBE_VERSION,
        "status": "failed",
        "phase": phase,
        "error": {
            "code": error_code,
            "message": message,
            "details": dict(details),
        },
        "capabilities": {"raw_cells_entered_host_context": False},
    }


def _load_json(path: str) -> Mapping[str, Any]:
    value = json.loads(Path(path).read_text(encoding="utf-8"))
    if not isinstance(value, dict):
        raise CapabilityError(f"expected JSON object: {path}")
    return value


def _write_receipt(path: str, receipt: Mapping[str, Any]) -> None:
    destination = Path(path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_bytes(canonical_json_bytes(receipt))


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser()
    subparsers = parser.add_subparsers(dest="command", required=True)

    verify = subparsers.add_parser("verify")
    verify.add_argument("--export", required=True)
    verify.add_argument("--metadata-before", required=True)
    verify.add_argument("--metadata-after", required=True)
    verify.add_argument("--cleanup-contract", required=True)
    verify.add_argument("--receipt", required=True)

    failure = subparsers.add_parser("record-failure")
    failure.add_argument("--phase", required=True)
    failure.add_argument("--error-code", required=True)
    failure.add_argument("--message", required=True)
    failure.add_argument("--details", required=True)
    failure.add_argument("--receipt", required=True)
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    arguments = _build_parser().parse_args(argv)
    if arguments.command == "verify":
        receipt = passing_receipt(
            _load_json(arguments.metadata_before),
            _load_json(arguments.metadata_after),
            validate_export(arguments.export),
            _load_json(arguments.cleanup_contract),
        )
    else:
        receipt = failure_receipt(
            phase=arguments.phase,
            error_code=arguments.error_code,
            message=arguments.message,
            details=_load_json(arguments.details),
        )
    _write_receipt(arguments.receipt, receipt)
    return 0 if receipt["status"] == "passed" else 8


if __name__ == "__main__":
    raise SystemExit(main())
